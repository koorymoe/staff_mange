#!/usr/bin/env python3
"""يحوّل إكسلات الجي بي اس لملف JSON واحد يقراه المستورد (cmd/importgps).

الاستخدام:
    python3 extract_gps_excel.py <بيانات_الجي_بي_اس.xlsx> <الشرائح.xlsx> > gps_data.json

ليش خطوتين (بايثون بعدين Go)؟ نفس أسلوب cmd/importproducts — ما نضيف مكتبة
إكسل لمشروع Go لأجل استيراد لمرة وحدة.

الربط بين الملفين: عمود «رقم ال GPS» بملف الزبائن = عمود SUBNO بملف الشرائح.
هذا هو المفتاح الي يخلينا نعرف أي شريحة بيد أي زبون.
"""
import json
import sys
import warnings
from datetime import datetime, date

import openpyxl

warnings.filterwarnings("ignore")

# الحالة بالإكسل → حالة الشريحة بالنظام
SIM_STATUS = {
    "مفعل": "IN_USE",
    "منتهي": "IN_USE",       # الاشتراك خلص بس الشريحة لسه بيد الزبون
    "تم الحرق": "BURNED",
    "غير موجود": "AVAILABLE",  # مو موجودة عند المشغل — نعتبرها متاحة
}

OPERATOR = {"زين": "ZAIN", "اسياسيل": "ASIACELL", "آسياسيل": "ASIACELL", "كورك": "KOREK"}


def clean(v):
    """يشيل المسافات ويحوّل الأرقام العشرية لنص صحيح (7834675993.0 → 7834675993)."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s or None


def iso(v):
    """تاريخ ISO — نرفض التواريخ غير المنطقية.

    بالإكسل صف تجريبي تاريخه 1906، ولو دخل النظام يطلع "صارله 44014 يوم"
    بقائمة الاتصالات ويشوّه الأرقام.
    """
    if isinstance(v, (datetime, date)):
        if not (2015 <= v.year <= 2100):
            return None
        return v.strftime("%Y-%m-%d")
    return None


def truthy(v):
    return v is True or (isinstance(v, str) and v.strip() in ("TRUE", "True", "نعم"))


def num(v):
    """بعض الخلايا الرقمية مخزونة كنص بالإكسل (التكاليف ورقم التبليغ) — نحوّلها."""
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).strip().replace(",", ""))
    except ValueError:
        return None


def main():
    if len(sys.argv) < 3:
        sys.exit("الاستخدام: extract_gps_excel.py <بيانات_GPS.xlsx> <الشرائح.xlsx>")

    # ── الشرائح ──
    ws = openpyxl.load_workbook(sys.argv[2], data_only=True)["Sheet1"]
    sims = []
    for r in range(2, ws.max_row + 1):
        sub = clean(ws.cell(r, 1).value)
        if not sub:
            continue
        raw_status = clean(ws.cell(r, 4).value)
        sims.append({
            "simNumber": sub,
            "iccid": clean(ws.cell(r, 2).value),
            "operator": OPERATOR.get(clean(ws.cell(r, 3).value) or "", "OTHER"),
            "status": SIM_STATUS.get(raw_status or "", "AVAILABLE"),
            "rawStatus": raw_status,
        })

    # ── الزبائن والأجهزة ── (العناوين بالصف ٣، البيانات من الصف ٤)
    ws = openpyxl.load_workbook(sys.argv[1], data_only=True)["بيانات الGPS"]
    customers = []
    for r in range(4, ws.max_row + 1):
        name = clean(ws.cell(r, 2).value)
        if not name:
            continue
        phone = clean(ws.cell(r, 3).value)
        if phone and not phone.startswith("0"):
            phone = "0" + phone
        customers.append({
            "fullName": name,
            "phone": phone,
            # رقم الجي بي اس هو نفسه SUBNO مال الشريحة — هذا مفتاح الربط
            "gpsNumber": clean(ws.cell(r, 4).value),
            "deviceImei": clean(ws.cell(r, 5).value),
            "subscriptionEnd": iso(ws.cell(r, 18).value) or iso(ws.cell(r, 6).value),
            "installedAt": iso(ws.cell(r, 17).value),
            "requestedAt": iso(ws.cell(r, 1).value),
            "installerName": clean(ws.cell(r, 13).value),
            "installCost": num(ws.cell(r, 16).value),
            "vehicleType": clean(ws.cell(r, 8).value),
            "installNote": clean(ws.cell(r, 7).value),
            "status": clean(ws.cell(r, 20).value),
            "notifyStage": num(ws.cell(r, 21).value),
            "notified1": truthy(ws.cell(r, 22).value),
            "notified2": truthy(ws.cell(r, 23).value),
            "notified40": truthy(ws.cell(r, 24).value),
            "simBurned": truthy(ws.cell(r, 25).value),
        })

    json.dump({"sims": sims, "customers": customers}, sys.stdout, ensure_ascii=False, indent=1)


if __name__ == "__main__":
    main()
